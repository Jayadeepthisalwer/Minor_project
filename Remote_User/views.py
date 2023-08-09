from django.db.models import Count
from django.shortcuts import render, redirect, get_object_or_404
import datetime
from django.db.models import Q
import openpyxl

# Create your views here.
from Remote_User.models import review_Model,ClientRegister_Model,productdetails_model,recommend_Model


def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:

            enter = ClientRegister_Model.objects.get(username=username, password=password)
            request.session["userid"] = enter.id

            return redirect('Add_Tweets')
        except:
            pass

    return render(request,'RUser/login.html')



def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:

        return render(request,'RUser/Register1.html')

def Add_Tweets(request):

    if "GET" == request.method:
        return render(request, 'RUser/Add_Tweets.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

            productdetails_model.objects.all().delete()
    recommend_Model.objects.all().delete()
    review_Model.objects.all().delete()

    for r in range(1, active_sheet.max_row+1):
        productdetails_model.objects.create(
        p_desc=active_sheet.cell(r, 1).value,
        names=active_sheet.cell(r, 2).value,
        sanalysis=active_sheet.cell(r, 3).value,
        senderstatus=active_sheet.cell(r, 4).value,
        ratings=active_sheet.cell(r, 5).value,
        likes=active_sheet.cell(r, 6).value,
        p_price=active_sheet.cell(r, 7).value,
        pcat=active_sheet.cell(r, 8).value,
        c_name=active_sheet.cell(r, 9).value,
        p_uses=active_sheet.cell(r, 10).value,
        dislikes=active_sheet.cell(r, 11).value,
        DT=active_sheet.cell(r, 12).value,
        topics="",
        uname=active_sheet.cell(r, 13).value

        )

    return render(request,'RUser/Add_Tweets.html', {"excel_data": excel_data})


def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})

def Review(request,pk):
    userid = request.session['userid']
    userObj = ClientRegister_Model.objects.get(id=userid)
    username = userObj.username

    objs = productdetails_model.objects.get(id=pk)
    tname = objs.names

    datetime_object = datetime.datetime.now()

    result = ''
    pos = []
    neg = []
    oth = []
    se = 'se'
    if request.method == "POST":
        uname = request.POST.get('uname')
        pname = request.POST.get('tname')
        feedback = request.POST.get('used')
        cmd = request.POST.get('review')

        if '#' in cmd:
            startingpoint = cmd.find('#')
            a = cmd[startingpoint:]
            endingPoint = a.find(' ')
            title = a[0:endingPoint]
            result = title[1:]
        # return redirect('')

        for f in cmd.split():
            if f in ('good', 'nice', 'better', 'best', 'excellent', 'extraordinary','beautiful', 'happy', 'won', 'love', 'greate',):
                pos.append(f)
            elif f in ('worst', 'waste', 'poor', 'error', 'imporve', 'bad', 'ridicules'):
                neg.append(f)
            else:
                oth.append(f)
        if len(pos) > len(neg):
            se = 'positive'
        elif len(neg) > len(pos):
            se = 'negative'
        else:
            se = 'neutral'
        review_Model.objects.create(uname=uname , ureview=cmd,sanalysis=se,dt=datetime_object,tname=pname ,feedback=feedback)

    return render(request,'RUser/Review.html', {'objc':username,'objc1':tname,'result': result, 'se': se})

def Recommend(request,pk):
    userid = request.session['userid']
    userObj = ClientRegister_Model.objects.get(id=userid)
    username = userObj.username

    objs = productdetails_model.objects.get(id=pk)
    appname = objs.names

    if request.method == "POST":
        username1 = request.POST.get('uname')
        pname = request.POST.get('pname')
        loc = request.POST.get('loc')
        usefull = request.POST.get('usefull')

        datetime_object = datetime.datetime.now()

        recommend_Model.objects.create( usefull=usefull, uname1=username1,loc=loc, dt=datetime_object, pname=pname)

    return render(request,'RUser/Recommend.html', {'objc':username,'objc1':appname})


def View_All_Tweet_Details(request):
    userid = request.session['userid']
    userObj = ClientRegister_Model.objects.get(id=userid)
    username = userObj.username


    obj = productdetails_model.objects.all().filter(~Q(uname=username))

    return render(request,'RUser/View_All_Tweet_Details.html',{'list_objects': obj})

def View_Tweet_Reviews(request):
    userid = request.session['userid']
    userObj = ClientRegister_Model.objects.get(id=userid)
    username = userObj.username

    obj = review_Model.objects.all()

    return render(request,'RUser/View_Tweet_Reviews.html',{'list_objects': obj})




def ratings(request,pk):
    vott1, vott, neg = 0, 0, 0
    objs = productdetails_model.objects.get(id=pk)
    unid = objs.id
    vot_count = productdetails_model.objects.all().filter(id=unid)
    for t in vot_count:
        vott = t.ratings
        vott1 = vott + 1
        obj = get_object_or_404(productdetails_model, id=unid)
        obj.ratings = vott1
        obj.save(update_fields=["ratings"])
        return redirect('View_All_Tweet_Details')

    return render(request,'RUser/ratings.html',{'objs':vott1})


def dislikes(request,pk):
    vott1, vott, neg = 0, 0, 0
    objs = productdetails_model.objects.get(id=pk)
    unid = objs.id
    vot_count = productdetails_model.objects.all().filter(id=unid)
    for t in vot_count:
        vott = t.dislikes
        vott1 = vott - 1
        obj = get_object_or_404(productdetails_model, id=unid)
        obj.dislikes = vott1
        obj.save(update_fields=["dislikes"])
        return redirect('View_All_Tweet_Details')
    return render(request,'RUser/dislikes.html',{'objs':vott1})

def ViewTrending(request):
    topic = productdetails_model.objects.values('topics').annotate(dcount=Count('topics')).order_by('-dcount')
    return render(request, 'RUser/ViewTrending.html', {'objects': topic})

def View_Tweet_Recommends(request):
    userid = request.session['userid']
    userObj = ClientRegister_Model.objects.get(id=userid)
    username = userObj.username

    obj = recommend_Model.objects.all().filter(~Q(uname1=username))

    return render(request, 'RUser/View_Tweet_Recommends.html', {'list_objects': obj})

def likes(request,pk):
    vott1, vott, neg = 0, 0, 0
    objs = productdetails_model.objects.get(id=pk)
    unid = objs.id
    vot_count = productdetails_model.objects.all().filter(id=unid)
    for t in vot_count:
        vott = t.likes
        vott1 = vott + 1
        obj = get_object_or_404(productdetails_model, id=unid)
        obj.likes = vott1
        obj.save(update_fields=["likes"])
        return redirect('View_All_Tweet_Details')

    return render(request,'RUser/likes.html',{'objs':vott1})

